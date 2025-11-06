from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for
import sqlite3
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

# ---------- DATABASE SETUP ----------
def init_db():
    conn = sqlite3.connect('queue.db')
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS queue (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    phone TEXT NOT NULL,
                    email TEXT,
                    latitude TEXT,
                    longitude TEXT
                )''')
    conn.commit()
    conn.close()

init_db()

# ---------- EXCEL SYNC ----------
def update_excel():
    db_conn = sqlite3.connect('queue.db')
    db_cursor = db_conn.cursor()
    db_cursor.execute("SELECT * FROM queue")
    data = db_cursor.fetchall()
    db_conn.close()

    file_path = "QueueData.xlsx"
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "Name", "Phone", "Email", "Latitude", "Longitude"])
    else:
        wb = load_workbook(file_path)
        ws = wb.active
        ws.delete_rows(2, ws.max_row)  # clear old data except headers

    for row in data:
        ws.append(row)

    wb.save(file_path)

# ---------- ROUTES ----------
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/register', methods=['POST'])
def register():
    data = request.get_json()
    conn = sqlite3.connect('queue.db')
    c = conn.cursor()
    c.execute("INSERT INTO queue (name, phone, email, latitude, longitude) VALUES (?, ?, ?, ?, ?)", 
              (data['name'], data['phone'], data['email'], data['latitude'], data['longitude']))
    conn.commit()
    queue_number = c.lastrowid
    conn.close()

    update_excel()  # Sync Excel after every registration

    return jsonify({"message": "Registration successful!", "queue_number": queue_number})

@app.route('/admin')
def admin():
    conn = sqlite3.connect('queue.db')
    c = conn.cursor()
    c.execute("SELECT * FROM queue")
    users = c.fetchall()
    conn.close()
    return render_template('admin.html', users=users)

@app.route('/delete/<int:user_id>', methods=['POST'])
def delete_user(user_id):
    conn = sqlite3.connect('queue.db')
    c = conn.cursor()
    
    # 1️⃣ Delete the selected user
    c.execute("DELETE FROM queue WHERE id = ?", (user_id,))
    conn.commit()
    
    # 2️⃣ Get remaining users ordered by current ID
    c.execute("SELECT name, phone, email, latitude, longitude FROM queue ORDER BY id ASC")
    remaining_users = c.fetchall()
    
    # 3️⃣ Drop and recreate the table (to reset IDs)
    c.execute("DROP TABLE queue")
    c.execute('''CREATE TABLE queue (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    phone TEXT NOT NULL,
                    email TEXT,
                    latitude TEXT,
                    longitude TEXT
                )''')
    
    # 4️⃣ Reinsert users (new IDs start from 1)
    for user in remaining_users:
        c.execute("INSERT INTO queue (name, phone, email, latitude, longitude) VALUES (?, ?, ?, ?, ?)", user)
    
    conn.commit()
    conn.close()
    
    # 5️⃣ Update Excel after reordering
    update_excel()
    
    return redirect(url_for('admin'))


@app.route('/export')
def export_excel():
    file_path = "QueueData.xlsx"
    if not os.path.exists(file_path):
        update_excel()
    return send_file(file_path, as_attachment=True)

@app.route('/queue')
def queue_page():
    conn = sqlite3.connect('queue.db')
    c = conn.cursor()
    c.execute("SELECT id, name, phone FROM queue")
    users = c.fetchall()
    conn.close()
    return render_template('queue.html', users=users)

# ---------- MAIN ----------
if __name__ == '__main__':
    app.run(debug=True)
